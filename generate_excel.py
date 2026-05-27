import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# ============================================================
# DATA: Pet-Friendly Condos in Bangkok grouped by 50 districts
# ============================================================

# All 50 Bangkok districts
DISTRICTS_TH = [
    "เขตพระนคร", "เขตดุสิต", "เขตหนองจอก", "เขตบางรัก", "เขตบางเขน",
    "เขตบางกะปิ", "เขตปทุมวัน", "เขตป้อมปราบศัตรูพ่าย", "เขตพระโขนง", "เขตมีนบุรี",
    "เขตลาดกระบัง", "เขตยานนาวา", "เขตสัมพันธวงศ์", "เขตพญาไท", "เขตธนบุรี",
    "เขตบางกอกใหญ่", "เขตห้วยขวาง", "เขตคลองสาน", "เขตตลิ่งชัน", "เขตบางกอกน้อย",
    "เขตบางขุนเทียน", "เขตภาษีเจริญ", "เขตหนองแขม", "เขตราษฎร์บูรณะ", "เขตบางพลัด",
    "เขตดินแดง", "เขตบึงกุ่ม", "เขตสาทร", "เขตบางซื่อ", "เขตจตุจักร",
    "เขตบางคอแหลม", "เขตประเวศ", "เขตคลองเตย", "เขตสวนหลวง", "เขตจอมทอง",
    "เขตดอนเมือง", "เขตราชเทวี", "เขตลาดพร้าว", "เขตวัฒนา", "เขตบางแค",
    "เขตหลักสี่", "เขตสายไหม", "เขตคันนายาว", "เขตสะพานสูง", "เขตวังทองหลาง",
    "เขตคลองสามวา", "เขตบางนา", "เขตทวีวัฒนา", "เขตทุ่งครุ", "เขตบางบอน", "เขตพระประแดง"
]

DISTRICTS_EN = [
    "Phra Nakhon", "Dusit", "Nong Chok", "Bang Rak", "Bang Khen",
    "Bang Kapi", "Pathum Wan", "Pom Prap Sattru Phai", "Phra Khanong", "Min Buri",
    "Lat Krabang", "Yan Nawa", "Samphanthawong", "Phaya Thai", "Thon Buri",
    "Bangkok Yai", "Huai Khwang", "Khlong San", "Taling Chan", "Bangkok Noi",
    "Bang Khun Thian", "Phasi Charoen", "Nong Khaem", "Rat Burana", "Bang Phlat",
    "Din Daeng", "Bueng Kum", "Sathon", "Bang Sue", "Chatuchak",
    "Bang Kho Laem", "Prawet", "Khlong Toei", "Suan Luang", "Chom Thong",
    "Don Mueang", "Ratchathewi", "Lat Phrao", "Watthana", "Bang Khae",
    "Lak Si", "Sai Mai", "Khan Na Yao", "Saphan Sung", "Wang Thonglang",
    "Khlong Sam Wa", "Bang Na", "Thawi Watthana", "Thung Khru", "Bang Bon", "Phra Pradaeng"
]

# Group district index by zone
DISTRICT_MAP = {}
for i, (th, en) in enumerate(zip(DISTRICTS_TH, DISTRICTS_EN)):
    DISTRICT_MAP[th] = i
    DISTRICT_MAP[en] = i
    DISTRICT_MAP[th.replace("เขต", "")] = i

# Condo data structure: name, district_idx, address, lat, lon, units, room_sizes, sale_price, rent_price, developer, phone, email
# price ranges for rent (THB/month): <10K, 10K-20K, 20K-50K, 50K-100K, >100K
# price ranges for sale (THB): <2M, 2M-5M, 5M-10M, 10M-20M, >20M

condos = [
    # ===== เขตวัฒนา (Watthana) - Sukhumvit core =====
    {
        "name": "Park Origin Thonglor",
        "name_th": "พาร์ค ออริจิ้น ทองหล่อ",
        "district": "เขตวัฒนา",
        "address": "ซอยทองหล่อ 10 ถนนสุขุมวิท 55 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7325, "lon": 100.5800,
        "units": 1182,
        "room_sizes": "30-97 ตร.ม.",
        "sale_price": "7,500,000 - 26,000,000",
        "rent_price": "26,000 - 120,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "The Monument Thonglor",
        "name_th": "เดอะ โมนูเมนต์ ทองหล่อ",
        "district": "เขตวัฒนา",
        "address": "998 ถนนทองหล่อ ซอยสุขุมวิท 55 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7340, "lon": 100.5820,
        "units": 127,
        "room_sizes": "125-295 ตร.ม.",
        "sale_price": "25,000,000 - 80,000,000",
        "rent_price": "90,000 - 200,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "สุนัข ≤15 กก., แมว, สัตว์ขนาดเล็ก"
    },
    {
        "name": "M Thonglor 10",
        "name_th": "เอ็ม ทองหล่อ 10",
        "district": "เขตวัฒนา",
        "address": "ซอยทองหล่อ 10 ถนนสุขุมวิท 55 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7310, "lon": 100.5780,
        "units": 160,
        "room_sizes": "35-144 ตร.ม.",
        "sale_price": "4,650,000 - 35,000,000",
        "rent_price": "20,000 - 60,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "The Strand Thonglor",
        "name_th": "เดอะ สแตรนด์ ทองหล่อ",
        "district": "เขตวัฒนา",
        "address": "ซอยทองหล่อ แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7330, "lon": 100.5810,
        "units": 100,
        "room_sizes": "80-200 ตร.ม.",
        "sale_price": "20,000,000 - 60,000,000",
        "rent_price": "150,000 - 250,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด (แจ้งล่วงหน้า)"
    },
    {
        "name": "Fullerton Sukhumvit",
        "name_th": "ฟูลเลอตัน สุขุมวิท",
        "district": "เขตวัฒนา",
        "address": "1219/2 ถนนสุขุมวิท แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7350, "lon": 100.5750,
        "units": 139,
        "room_sizes": "105-152 ตร.ม.",
        "sale_price": "14,500,000 - 34,900,000",
        "rent_price": "57,000 - 100,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Maru Ekkamai 2",
        "name_th": "มารุ เอกมัย 2",
        "district": "เขตวัฒนา",
        "address": "70-5 ถนนสุขุมวิท 63 (เอกมัย) แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7240, "lon": 100.5860,
        "units": 333,
        "room_sizes": "28-65 ตร.ม.",
        "sale_price": "4,500,000 - 8,500,000",
        "rent_price": "22,000 - 85,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS เอกมัย",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "The Madison",
        "name_th": "เดอะ เมดิสัน",
        "district": "เขตวัฒนา",
        "address": "737 ถนนสุขุมวิท แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7360, "lon": 100.5770,
        "units": 177,
        "room_sizes": "80-300 ตร.ม.",
        "sale_price": "24,500,000 - 218,000,000",
        "rent_price": "60,000 - 200,000",
        "developer": "Raimon Land",
        "phone": "02-661-7777",
        "email": "info@raimonland.com",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Scope Promsri",
        "name_th": "สโคป พร้อมศรี",
        "district": "เขตวัฒนา",
        "address": "345 ถนนสุขุมวิท 49 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7345, "lon": 100.5790,
        "units": 150,
        "room_sizes": "40-120 ตร.ม.",
        "sale_price": "6,900,000 - 20,000,000",
        "rent_price": "30,000 - 80,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "Muniq Sukhumvit 23",
        "name_th": "มิวนีค สุขุมวิท 23",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 23 แขวงคลองเตยเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7390, "lon": 100.5620,
        "units": 201,
        "room_sizes": "34.74-191.11 ตร.ม.",
        "sale_price": "5,400,000 - 80,000,000",
        "rent_price": "35,000 - 130,000",
        "developer": "Major Development",
        "phone": "087-778-1111",
        "email": "info@mde.co.th",
        "bts": "BTS อโศก / MRT สุขุมวิท",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "The Estelle Phrom Phong",
        "name_th": "ดิ เอสเทลล์ พร้อมพงษ์",
        "district": "เขตวัฒนา",
        "address": "131 ซอยสุขุมวิท 26 แขวงคลองตัน เขตคลองเตย กรุงเทพฯ",
        "lat": 13.7280, "lon": 100.5670,
        "units": 157,
        "room_sizes": "48-99.5 ตร.ม.",
        "sale_price": "15,000,000 - 50,000,000",
        "rent_price": "50,000 - 150,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Via 61",
        "name_th": "เวีย 61",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 61 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7250, "lon": 100.5900,
        "units": 150,
        "room_sizes": "50-187 ตร.ม.",
        "sale_price": "10,900,000 - 30,000,000",
        "rent_price": "35,000 - 90,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS เอกมัย",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "W8 Thonglor 25",
        "name_th": "ดับบลิว เอท ทองหล่อ 25",
        "district": "เขตวัฒนา",
        "address": "ซอยทองหล่อ 25 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7290, "lon": 100.5830,
        "units": 80,
        "room_sizes": "35-80 ตร.ม.",
        "sale_price": "6,000,000 - 18,000,000",
        "rent_price": "25,000 - 70,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Baan Ananda",
        "name_th": "บ้าน อนันดา",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 63 (เอกมัย) แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7230, "lon": 100.5870,
        "units": 90,
        "room_sizes": "70-150 ตร.ม.",
        "sale_price": "12,000,000 - 30,000,000",
        "rent_price": "150,000+",
        "developer": "Ananda",
        "phone": "02-123-4567",
        "email": "info@ananda.co.th",
        "bts": "BTS เอกมัย",
        "pet_policy": "อนุญาตแมวและสุนัข (ต้องแจ้งล่วงหน้า)"
    },
    {
        "name": "Le Nice Ekamai",
        "name_th": "เลอ นิส เอกมัย",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 63 (เอกมัย) แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7220, "lon": 100.5880,
        "units": 120,
        "room_sizes": "30-60 ตร.ม.",
        "sale_price": "4,000,000 - 8,000,000",
        "rent_price": "25,000 - 45,000",
        "developer": "LPN Development",
        "phone": "02-123-4567",
        "email": "info@lpn.co.th",
        "bts": "BTS เอกมัย",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "Eden Ekkamai",
        "name_th": "อีเดน เอกมัย",
        "district": "เขตวัฒนา",
        "address": "ซอยเอกมัย แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7245, "lon": 100.5875,
        "units": 80,
        "room_sizes": "35-75 ตร.ม.",
        "sale_price": "5,000,000 - 12,000,000",
        "rent_price": "20,000 - 50,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS เอกมัย",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Nivati Thonglor 23",
        "name_th": "นิวาติ ทองหล่อ 23",
        "district": "เขตวัฒนา",
        "address": "ซอยทองหล่อ 23 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7300, "lon": 100.5840,
        "units": 60,
        "room_sizes": "45-120 ตร.ม.",
        "sale_price": "8,000,000 - 25,000,000",
        "rent_price": "35,000 - 80,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "La Citta Delre Thonglor 16",
        "name_th": "ลา ซิตต้า เดลเร ทองหล่อ 16",
        "district": "เขตวัฒนา",
        "address": "ซอยทองหล่อ 16 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7315, "lon": 100.5825,
        "units": 70,
        "room_sizes": "40-100 ตร.ม.",
        "sale_price": "7,000,000 - 20,000,000",
        "rent_price": "30,000 - 65,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Downtown 49",
        "name_th": "ดาวทาวน์ โฟร์ตี้ไนน์",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 49 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7335, "lon": 100.5785,
        "units": 95,
        "room_sizes": "35-80 ตร.ม.",
        "sale_price": "4,500,000 - 12,000,000",
        "rent_price": "18,000 - 45,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Reference Ekkamai",
        "name_th": "เรฟเฟอเรนซ์ เอกมัย",
        "district": "เขตวัฒนา",
        "address": "ซอยเอกมัย แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7260, "lon": 100.5860,
        "units": 110,
        "room_sizes": "30-65 ตร.ม.",
        "sale_price": "3,500,000 - 8,000,000",
        "rent_price": "18,000 - 40,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS เอกมัย",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Valles Haus",
        "name_th": "วาลเลส เฮาส์",
        "district": "เขตวัฒนา",
        "address": "55 ถนนฮะโบ แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.7200, "lon": 100.5950,
        "units": 400,
        "room_sizes": "33.75-90.25 ตร.ม.",
        "sale_price": "4,690,000 - 18,211,000",
        "rent_price": "15,000 - 40,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS อ่อนนุช",
        "pet_policy": "อนุญาตสัตว์เลี้ยง (โครงการ Pet-Welcome)"
    },
    {
        "name": "Pynn Pridi 20",
        "name_th": "พินน์ ปรีดี 20",
        "district": "เขตวัฒนา",
        "address": "ซอยปรีดีพนมยงค์ 20 แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7210, "lon": 100.5940,
        "units": 200,
        "room_sizes": "33.5-75.5 ตร.ม.",
        "sale_price": "6,490,000 - 15,000,000",
        "rent_price": "25,000 - 55,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS พระโขนง",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "Maestro 39",
        "name_th": "มาเอสโตร 39",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 39 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7340, "lon": 100.5760,
        "units": 150,
        "room_sizes": "35-80 ตร.ม.",
        "sale_price": "4,500,000 - 12,000,000",
        "rent_price": "25,000 - 50,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "Scope Thonglor",
        "name_th": "สโคป ทองหล่อ",
        "district": "เขตวัฒนา",
        "address": "ซอยทองหล่อ 13 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.731, "lon": 100.578,
        "units": 160,
        "room_sizes": "45-130 ตร.ม.",
        "sale_price": "12,000,000 - 40,000,000",
        "rent_price": "40,000 - 120,000",
        "developer": "SC Asset",
        "phone": "02-119-0000",
        "email": "info@scasset.com",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "Quintara MHy'ZEN Phrom Phong",
        "name_th": "ควินทาร่า มายเซน พร้อมพงษ์",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 39 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.734, "lon": 100.568,
        "units": 230,
        "room_sizes": "30-75 ตร.ม.",
        "sale_price": "7,000,000 - 18,000,000",
        "rent_price": "25,000 - 70,000",
        "developer": "Eastern Star",
        "phone": "02-118-0000",
        "email": "info@easternstar.co.th",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Muniq Sukhumvit 23",
        "name_th": "มิวนีค สุขุมวิท 23",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 23 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.739, "lon": 100.562,
        "units": 180,
        "room_sizes": "35-75 ตร.ม.",
        "sale_price": "8,000,000 - 20,000,000",
        "rent_price": "35,000 - 80,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS อโศก / MRT สุขุมวิท",
        "pet_policy": "Pet-Friendly"
    },
    {
        "name": "Somerset Ekamai Bangkok",
        "name_th": "ซัมเมอร์เซ็ต เอกมัย",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 63 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.728, "lon": 100.585,
        "units": 180,
        "room_sizes": "35-120 ตร.ม.",
        "sale_price": "5,000,000 - 15,000,000",
        "rent_price": "30,000 - 80,000",
        "developer": "Somerset",
        "phone": "02-123-4567",
        "email": "info@somerset.com",
        "bts": "BTS เอกมัย",
        "pet_policy": "อนุญาตสัตว์เลี้ยง (Serviced Apartment)"
    },
    {
        "name": "Marque Sukhumvit",
        "name_th": "มาร์ค สุขุมวิท",
        "district": "เขตวัฒนา",
        "address": "ซอยสุขุมวิท 41 แขวงคลองตันเหนือ เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.736, "lon": 100.572,
        "units": 220,
        "room_sizes": "40-120 ตร.ม.",
        "sale_price": "12,000,000 - 35,000,000",
        "rent_price": "50,000 - 150,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS พร้อมพงษ์",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "Valles Haus",
        "name_th": "วัลเลส เฮาส์",
        "district": "เขตวัฒนา",
        "address": "สุขุมวิท แขวงพระโขนง เขตวัฒนา กรุงเทพฯ 10110",
        "lat": 13.716, "lon": 100.592,
        "units": 200,
        "room_sizes": "28-65 ตร.ม.",
        "sale_price": "4,690,000 - 12,000,000",
        "rent_price": "18,000 - 50,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS ทองหล่อ / BTS พระโขนง",
        "pet_policy": "Pet-Friendly Project (เปิดใหม่ 2027)"
    },

    # ===== เขตคลองเตย (Khlong Toei) =====
    {
        "name": "Ideo Morph 38",
        "name_th": "ไอดีโอ มอร์ฟ 38",
        "district": "เขตคลองเตย",
        "address": "ซานติสุข ซอยสุขุมวิท 38 แขวงพระโขนง เขตคลองเตย กรุงเทพฯ 10110",
        "lat": 13.72098, "lon": 100.57930,
        "units": 162,
        "room_sizes": "23-45 ตร.ม.",
        "sale_price": "6,900,000 - 9,000,000",
        "rent_price": "28,000 - 36,000",
        "developer": "Ananda",
        "phone": "02-123-4567",
        "email": "info@ananda.co.th",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Aguston Sukhumvit 22",
        "name_th": "ออกัสตัน สุขุมวิท 22",
        "district": "เขตคลองเตย",
        "address": "222/23 ซอยสุขุมวิท 22 แขวงคลองเตย เขตคลองเตย กรุงเทพฯ 10110",
        "lat": 13.72514, "lon": 100.56488,
        "units": 269,
        "room_sizes": "50-280 ตร.ม.",
        "sale_price": "6,850,000 - 25,000,000",
        "rent_price": "30,000 - 150,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS อโศก / MRT สุขุมวิท",
        "pet_policy": "อนุญาตแมวและสุนัข (ต้องแจ้งล่วงหน้า)"
    },
    {
        "name": "Ashton Morph 38",
        "name_th": "แอชตัน มอร์ฟ 38",
        "district": "เขตคลองเตย",
        "address": "ซอยสุขุมวิท 38 แขวงพระโขนง เขตคลองเตย กรุงเทพฯ 10110",
        "lat": 13.7215, "lon": 100.5800,
        "units": 199,
        "room_sizes": "52-295 ตร.ม.",
        "sale_price": "6,200,000 - 37,450,000",
        "rent_price": "29,000 - 60,000",
        "developer": "Ananda",
        "phone": "02-123-4567",
        "email": "info@ananda.co.th",
        "bts": "BTS ทองหล่อ",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Cosmo Villa",
        "name_th": "คอสโม่ วิลล่า",
        "district": "เขตคลองเตย",
        "address": "ซอยสุขุมวิท 12 แขวงคลองเตย เขตคลองเตย กรุงเทพฯ",
        "lat": 13.7370, "lon": 100.5560,
        "units": 60,
        "room_sizes": "30-80 ตร.ม.",
        "sale_price": "3,500,000 - 10,000,000",
        "rent_price": "55,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS อโศก",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก (ต้องแจ้งล่วงหน้า)"
    },
    {
        "name": "Maestro 01 Sathorn-Yenakat",
        "name_th": "มาเอสโตร 01 สาทร-เย็นอากาศ",
        "district": "เขตคลองเตย",
        "address": "ถนนเย็นอากาศ แขวงทุ่งมหาเมฆ เขตสาทร กรุงเทพฯ 10120",
        "lat": 13.7150, "lon": 100.5500,
        "units": 100,
        "room_sizes": "33.56-182.18 ตร.ม.",
        "sale_price": "3,000,000 - 10,000,000",
        "rent_price": "20,000 - 45,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT คลองเตย",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Ficus Lane",
        "name_th": "ไฟคัส เลน",
        "district": "เขตคลองเตย",
        "address": "ซอยสุขุมวิท 44/1 แขวงพระโขนง เขตคลองเตย กรุงเทพฯ",
        "lat": 13.71388, "lon": 100.58929,
        "units": 70,
        "room_sizes": "50-472 ตร.ม.",
        "sale_price": "18,500,000 - 99,000,000",
        "rent_price": "30,000 - 235,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS พระโขนง",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "The Pillar",
        "name_th": "เดอะ พิลลาร์",
        "district": "เขตคลองเตย",
        "address": "ซอยสุขุมวิท 44 แขวงพระโขนง เขตคลองเตย กรุงเทพฯ",
        "lat": 13.7140, "lon": 100.5880,
        "units": 80,
        "room_sizes": "40-100 ตร.ม.",
        "sale_price": "6,000,000 - 18,000,000",
        "rent_price": "25,000 - 60,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS พระโขนง",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตปทุมวัน (Pathum Wan) =====
    {
        "name": "Muniq Langsuan",
        "name_th": "มิวนีค หลังสวน",
        "district": "เขตปทุมวัน",
        "address": "ซอยต้นสน แขวงลุมพินี เขตปทุมวัน กรุงเทพฯ 10330",
        "lat": 13.7420, "lon": 100.5420,
        "units": 166,
        "room_sizes": "50-179 ตร.ม.",
        "sale_price": "15,000,000 - 160,000,000",
        "rent_price": "45,000 - 450,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS ชิดลม",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Scope Langsuan",
        "name_th": "สโคป หลังสวน",
        "district": "เขตปทุมวัน",
        "address": "ซอยหลังสวน แขวงลุมพินี เขตปทุมวัน กรุงเทพฯ 10330",
        "lat": 13.7410, "lon": 100.5430,
        "units": 157,
        "room_sizes": "82-164 ตร.ม.",
        "sale_price": "35,000,000 - 690,000,000",
        "rent_price": "129,999 - 500,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS ราชดำริ",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "Maru Chula",
        "name_th": "มารุ จุฬา",
        "district": "เขตปทุมวัน",
        "address": "117 ถนนเจริญเมือง แขวงรองเมือง เขตปทุมวัน กรุงเทพฯ 10330",
        "lat": 13.7380, "lon": 100.5280,
        "units": 220,
        "room_sizes": "30-60 ตร.ม.",
        "sale_price": "3,950,000 - 7,650,000",
        "rent_price": "15,000 - 35,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT สามย่าน",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Grand Langsuan",
        "name_th": "แกรนด์ หลังสวน",
        "district": "เขตปทุมวัน",
        "address": "ถนนหลังสวน แขวงลุมพินี เขตปทุมวัน กรุงเทพฯ",
        "lat": 13.7415, "lon": 100.5440,
        "units": 200,
        "room_sizes": "80-200 ตร.ม.",
        "sale_price": "20,000,000 - 80,000,000",
        "rent_price": "100,000 - 200,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS ชิดลม",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Tonson One Residence",
        "name_th": "ต้นสน วัน เรสซิเดนซ์",
        "district": "เขตปทุมวัน",
        "address": "ซอยต้นสน แขวงลุมพินี เขตปทุมวัน กรุงเทพฯ",
        "lat": 13.7425, "lon": 100.5410,
        "units": 50,
        "room_sizes": "80-250 ตร.ม.",
        "sale_price": "30,000,000 - 100,000,000",
        "rent_price": "250,000 - 500,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS ชิดลม",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "Manhattan Chidlom",
        "name_th": "แมนฮัตตัน ชิดลม",
        "district": "เขตปทุมวัน",
        "address": "ถนนชิดลม แขวงลุมพินี เขตปทุมวัน กรุงเทพฯ",
        "lat": 13.7440, "lon": 100.5440,
        "units": 100,
        "room_sizes": "35-80 ตร.ม.",
        "sale_price": "8,000,000 - 20,000,000",
        "rent_price": "25,000 - 55,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS ชิดลม",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Maestro 02 Ruamrudee",
        "name_th": "มาเอสโตร 02 ร่วมฤดี",
        "district": "เขตปทุมวัน",
        "address": "ซอยร่วมฤดี แขวงลุมพินี เขตปทุมวัน กรุงเทพฯ",
        "lat": 13.7400, "lon": 100.5480,
        "units": 80,
        "room_sizes": "35-70 ตร.ม.",
        "sale_price": "4,500,000 - 10,000,000",
        "rent_price": "20,000 - 40,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS เพลินจิต",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "98 Wireless",
        "name_th": "ไนน์ตี้เอท ไวร์เลส",
        "district": "เขตปทุมวัน",
        "address": "ถนนวิทยุ แขวงลุมพินี เขตปทุมวัน กรุงเทพฯ 10330",
        "lat": 13.740, "lon": 100.541,
        "units": 200,
        "room_sizes": "100-800 ตร.ม.",
        "sale_price": "30,000,000 - 200,000,000",
        "rent_price": "250,000 - 800,000",
        "developer": "Raimon Land",
        "phone": "02-168-9999",
        "email": "info@raimonland.com",
        "bts": "BTS เพลินจิต",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด (Luxury Residence)"
    },

    # ===== เขตบางรัก (Bang Rak) =====
    {
        "name": "M Silom",
        "name_th": "เอ็ม สีลม",
        "district": "เขตบางรัก",
        "address": "ถนนสีลม แขวงสุริยวงศ์ เขตบางรัก กรุงเทพฯ 10500",
        "lat": 13.7280, "lon": 100.5200,
        "units": 200,
        "room_sizes": "30-80 ตร.ม.",
        "sale_price": "5,000,000 - 15,000,000",
        "rent_price": "39,000 - 80,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS ช่องนนทรี",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "The Ritz-Carlton Residences Bangkok",
        "name_th": "เดอะ ริทซ์-คาร์ลตัน เรสซิเดนซ์ กรุงเทพฯ",
        "district": "เขตบางรัก",
        "address": "อาคารมหานคร ถนนพระราม 1 แขวงสีลม เขตบางรัก กรุงเทพฯ",
        "lat": 13.7240, "lon": 100.5280,
        "units": 200,
        "room_sizes": "80-600 ตร.ม.",
        "sale_price": "50,000,000 - 300,000,000",
        "rent_price": "450,000 - 1,500,000",
        "developer": "Pace Development",
        "phone": "02-123-4567",
        "email": "info@mahakhon.com",
        "bts": "BTS ช่องนนทรี",
        "pet_policy": "อนุญาตสัตว์เลี้ยง (แจ้งกรณีสัตว์ใหญ่)"
    },
    {
        "name": "The Royal Saladaeng",
        "name_th": "เดอะ รอยัล ศาลาแดง",
        "district": "เขตบางรัก",
        "address": "ถนนศาลาแดง แขวงสีลม เขตบางรัก กรุงเทพฯ",
        "lat": 13.7260, "lon": 100.5350,
        "units": 80,
        "room_sizes": "40-120 ตร.ม.",
        "sale_price": "8,000,000 - 30,000,000",
        "rent_price": "40,000 - 120,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS ศาลาแดง / MRT สีลม",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตสาทร (Sathon) =====
    {
        "name": "Tait 12",
        "name_th": "เทตต์ ทเวลฟ์",
        "district": "เขตสาทร",
        "address": "ซอยเซนต์หลุยส์ 2 แขวงทุ่งมหาเมฆ เขตสาทร กรุงเทพฯ 10120",
        "lat": 13.7190, "lon": 100.5350,
        "units": 100,
        "room_sizes": "45-120 ตร.ม.",
        "sale_price": "8,000,000 - 25,000,000",
        "rent_price": "40,000 - 90,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS เซนต์หลุยส์",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "Four Seasons Private Residence Bangkok",
        "name_th": "โฟร์ซีซั่นส์ ไพรเวท เรสซิเดนซ์ กรุงเทพฯ",
        "district": "เขตสาทร",
        "address": "ริมแม่น้ำเจ้าพระยา ถนนเจริญกรุง แขวงยานนาวา เขตสาทร กรุงเทพฯ",
        "lat": 13.7100, "lon": 100.5100,
        "units": 200,
        "room_sizes": "100-500 ตร.ม.",
        "sale_price": "40,000,000 - 300,000,000",
        "rent_price": "350,000 - 800,000",
        "developer": "Country Group Development",
        "phone": "02-123-4567",
        "email": "info@cgd.co.th",
        "bts": "BTS กรุงธนบุรี",
        "pet_policy": "Pet-friendly (มัดจำ 3 เดือน)"
    },
    {
        "name": "Collezio Sathorn-Pipat",
        "name_th": "คอลเลซิโอ สาทร-พิพัฒน์",
        "district": "เขตสาทร",
        "address": "ซอยพิพัฒน์ แขวงสีลม เขตบางรัก กรุงเทพฯ",
        "lat": 13.7200, "lon": 100.5300,
        "units": 80,
        "room_sizes": "30-70 ตร.ม.",
        "sale_price": "4,000,000 - 10,000,000",
        "rent_price": "35,000 - 60,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS ช่องนนทรี",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "The Peony",
        "name_th": "เดอะ พิโอนี่",
        "district": "เขตสาทร",
        "address": "ถนนสาทร แขวงยานนาวา เขตสาทร กรุงเทพฯ",
        "lat": 13.7140, "lon": 100.5250,
        "units": 60,
        "room_sizes": "35-80 ตร.ม.",
        "sale_price": "3,500,000 - 10,000,000",
        "rent_price": "20,000 - 45,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS เซนต์หลุยส์",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Maestro 01 Sathorn-Yenakat",
        "name_th": "มาเอสโตร 01 สาทร-เย็นอากาศ",
        "district": "เขตสาทร",
        "address": "ถนนเย็นอากาศ แขวงทุ่งมหาเมฆ เขตสาทร กรุงเทพฯ 10120",
        "lat": 13.7155, "lon": 100.5480,
        "units": 100,
        "room_sizes": "33.56-182.18 ตร.ม.",
        "sale_price": "3,000,000 - 10,000,000",
        "rent_price": "20,000 - 45,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT คลองเตย",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตพญาไท (Phaya Thai) =====
    {
        "name": "M Phayathai",
        "name_th": "เอ็ม พญาไท",
        "district": "เขตพญาไท",
        "address": "ถนนพญาไท แขวงสามเสนใน เขตพญาไท กรุงเทพฯ 10400",
        "lat": 13.7620, "lon": 100.5390,
        "units": 200,
        "room_sizes": "30-80 ตร.ม.",
        "sale_price": "8,000,000 - 29,000,000",
        "rent_price": "25,000 - 60,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS อนุสาวรีย์ชัย",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Marquis Phayathai",
        "name_th": "มาร์ควิส พญาไท",
        "district": "เขตพญาไท",
        "address": "ถนนพญาไท แขวงถนนพญาไท เขตราชเทวี กรุงเทพฯ 10400",
        "lat": 13.7600, "lon": 100.5400,
        "units": 180,
        "room_sizes": "45-150 ตร.ม.",
        "sale_price": "15,000,000 - 50,000,000",
        "rent_price": "45,000 - 120,000",
        "developer": "Major Development",
        "phone": "1266",
        "email": "info@mde.co.th",
        "bts": "BTS พญาไท",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "The Monument Sanampao",
        "name_th": "เดอะ โมนูเมนต์ สนามเป้า",
        "district": "เขตพญาไท",
        "address": "พหลโยธิน แขวงสามเสนใน เขตพญาไท กรุงเทพฯ 10400",
        "lat": 13.776, "lon": 100.543,
        "units": 130,
        "room_sizes": "45-120 ตร.ม.",
        "sale_price": "12,000,000 - 40,000,000",
        "rent_price": "60,000 - 160,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS สนามเป้า",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "Baxtor Paholyothin 14",
        "name_th": "แบ็กซ์เตอร์ พหลโยธิน 14",
        "district": "เขตพญาไท",
        "address": "พหลโยธิน 14 แขวงสามเสนใน เขตพญาไท กรุงเทพฯ 10400",
        "lat": 13.772, "lon": 100.543,
        "units": 150,
        "room_sizes": "25-55 ตร.ม.",
        "sale_price": "3,500,000 - 7,000,000",
        "rent_price": "15,000 - 35,000",
        "developer": "Property Perfect",
        "phone": "02-123-4567",
        "email": "info@propertyperfect.com",
        "bts": "BTS สะพานควาย",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },

    # ===== เขตราชเทวี (Ratchathewi) =====
    {
        "name": "Conner Ratchathewi",
        "name_th": "คอนเนอร์ ราชเทวี",
        "district": "เขตราชเทวี",
        "address": "ถนนพญาไท แขวงถนนเพชรบุรี เขตราชเทวี กรุงเทพฯ",
        "lat": 13.7530, "lon": 100.5360,
        "units": 150,
        "room_sizes": "35-90 ตร.ม.",
        "sale_price": "8,000,000 - 20,000,000",
        "rent_price": "35,000 - 70,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS ราชเทวี",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Maestro 12 Ratchathewi",
        "name_th": "มาเอสโตร 12 ราชเทวี",
        "district": "เขตราชเทวี",
        "address": "ถนนพญาไท แขวงถนนพญาไท เขตราชเทวี กรุงเทพฯ",
        "lat": 13.7550, "lon": 100.5350,
        "units": 100,
        "room_sizes": "28-60 ตร.ม.",
        "sale_price": "3,500,000 - 8,000,000",
        "rent_price": "18,000 - 35,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS ราชเทวี",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "Maestro 14 Siam-Ratchathewi",
        "name_th": "มาเอสโตร 14 สยาม-ราชเทวี",
        "district": "เขตราชเทวี",
        "address": "ถนนพญาไท แขวงถนนพญาไท เขตราชเทวี กรุงเทพฯ",
        "lat": 13.7540, "lon": 100.5370,
        "units": 100,
        "room_sizes": "28-55 ตร.ม.",
        "sale_price": "3,500,000 - 7,000,000",
        "rent_price": "18,000 - 30,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS สยาม",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Maestro 07 Victory Monument",
        "name_th": "มาเอสโตร 07 อนุสาวรีย์ชัย",
        "district": "เขตราชเทวี",
        "address": "ถนนราชวิถี แขวงถนนพญาไท เขตราชเทวี กรุงเทพฯ",
        "lat": 13.7650, "lon": 100.5400,
        "units": 120,
        "room_sizes": "28-60 ตร.ม.",
        "sale_price": "3,500,000 - 8,000,000",
        "rent_price": "20,000 - 40,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS อนุสาวรีย์ชัย",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกขนาด"
    },
    {
        "name": "Chewathai Residence Asoke",
        "name_th": "ชีวาทัย เรสซิเดนซ์ อโศก",
        "district": "เขตราชเทวี",
        "address": "อโศก-ดินแดง แขวงมักกะสัน เขตราชเทวี กรุงเทพฯ 10400",
        "lat": 13.756, "lon": 100.560,
        "units": 280,
        "room_sizes": "28-65 ตร.ม.",
        "sale_price": "4,000,000 - 10,000,000",
        "rent_price": "18,000 - 45,000",
        "developer": "Chewathai",
        "phone": "02-123-4567",
        "email": "info@chewathai.com",
        "bts": "MRT เพชรบุรี",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },

    # ===== เขตห้วยขวาง (Huai Khwang) =====
    {
        "name": "Nue Epic Asok-Rama 9",
        "name_th": "นิว เอปิค อโศก-พระราม 9",
        "district": "เขตห้วยขวาง",
        "address": "ถนนรัชดาภิเษก แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.7500, "lon": 100.5700,
        "units": 400,
        "room_sizes": "25-50 ตร.ม.",
        "sale_price": "3,500,000 - 7,000,000",
        "rent_price": "15,000 - 35,000",
        "developer": "Noble Development",
        "phone": "02-123-4567",
        "email": "info@noble.com",
        "bts": "MRT พระราม 9",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก (แจ้งล่วงหน้า)"
    },
    {
        "name": "COBE Ratchada-Rama 9",
        "name_th": "โค้บบ์ รัชดา-พระราม 9",
        "district": "เขตห้วยขวาง",
        "address": "ถนนรัชดาภิเษก แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.7550, "lon": 100.5750,
        "units": 300,
        "room_sizes": "25-55 ตร.ม.",
        "sale_price": "3,000,000 - 7,000,000",
        "rent_price": "13,000 - 30,000",
        "developer": "SC Asset",
        "phone": "02-123-4567",
        "email": "info@scasset.com",
        "bts": "MRT พระราม 9",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Maestro 03 Ratchada-Rama 9",
        "name_th": "มาเอสโตร 03 รัชดา-พระราม 9",
        "district": "เขตห้วยขวาง",
        "address": "ถนนรัชดาภิเษก แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.7520, "lon": 100.5720,
        "units": 200,
        "room_sizes": "28-65 ตร.ม.",
        "sale_price": "2,800,000 - 6,500,000",
        "rent_price": "15,900 - 35,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT พระราม 9",
        "pet_policy": "ไม่เกิน 15 กก./ตัว ไม่เกิน 1-2 ตัว"
    },
    {
        "name": "Maestro 19 Ratchada 19",
        "name_th": "มาเอสโตร 19 รัชดา 19",
        "district": "เขตห้วยขวาง",
        "address": "ซอยรัชดา 19 แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.7580, "lon": 100.5730,
        "units": 150,
        "room_sizes": "28-55 ตร.ม.",
        "sale_price": "2,500,000 - 5,500,000",
        "rent_price": "15,000 - 28,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT รัชดาภิเษก",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "The Base Ratchada 19",
        "name_th": "เดอะ เบส รัชดา 19",
        "district": "เขตห้วยขวาง",
        "address": "ซอยรัชดา 19 แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.7570, "lon": 100.5740,
        "units": 200,
        "room_sizes": "23.25-35.25 ตร.ม.",
        "sale_price": "2,490,000 - 4,500,000",
        "rent_price": "10,000 - 22,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "MRT รัชดาภิเษก",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Cybiq Rama 9",
        "name_th": "ไซบิค พระราม 9",
        "district": "เขตห้วยขวาง",
        "address": "ถนนพระราม 9 แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.7480, "lon": 100.5650,
        "units": 180,
        "room_sizes": "25-50 ตร.ม.",
        "sale_price": "2,500,000 - 5,000,000",
        "rent_price": "13,000 - 25,000",
        "developer": "Ananda",
        "phone": "02-123-4567",
        "email": "info@ananda.co.th",
        "bts": "MRT พระราม 9",
        "pet_policy": "Pet-friendly"
    },
    {
        "name": "Rise Rama 9",
        "name_th": "ไรซ์ พระราม 9",
        "district": "เขตห้วยขวาง",
        "address": "ถนนพระราม 9 แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.7490, "lon": 100.5670,
        "units": 250,
        "room_sizes": "22-45 ตร.ม.",
        "sale_price": "1,800,000 - 4,000,000",
        "rent_price": "9,000 - 20,000",
        "developer": "AP Thailand",
        "phone": "02-123-4567",
        "email": "info@apthai.com",
        "bts": "MRT พระราม 9",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Chapter One ECO Ratchada-Huaikwang",
        "name_th": "แชปเตอร์ วัน อีโค รัชดา-ห้วยขวาง",
        "district": "เขตห้วยขวาง",
        "address": "รัชดาภิเษก แขวงห้วยขวาง เขตห้วยขวาง กรุงเทพฯ",
        "lat": 13.775, "lon": 100.571,
        "units": 800,
        "room_sizes": "26-50 ตร.ม.",
        "sale_price": "2,800,000 - 5,500,000",
        "rent_price": "15,000 - 30,000",
        "developer": "Pruksa",
        "phone": "02-123-4567",
        "email": "info@pruksa.com",
        "bts": "MRT ห้วยขวาง",
        "pet_policy": "มี Pet Zone, อนุญาตแมวและสุนัขขนาดเล็ก"
    },

    # ===== เขตดินแดง (Din Daeng) =====
    {
        "name": "The Base Ratchada 19",
        "name_th": "เดอะ เบส รัชดา 19",
        "district": "เขตดินแดง",
        "address": "ถนนรัชดาภิเษก แขวงดินแดง เขตดินแดง กรุงเทพฯ",
        "lat": 13.7590, "lon": 100.5730,
        "units": 200,
        "room_sizes": "23.25-35.25 ตร.ม.",
        "sale_price": "2,490,000 - 4,500,000",
        "rent_price": "10,000 - 22,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "MRT รัชดาภิเษก",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตจตุจักร (Chatuchak) =====
    {
        "name": "M Jatujak",
        "name_th": "เอ็ม จตุจักร",
        "district": "เขตจตุจักร",
        "address": "ถนนพหลโยธิน แขวงจตุจักร เขตจตุจักร กรุงเทพฯ 10900",
        "lat": 13.8150, "lon": 100.5620,
        "units": 800,
        "room_sizes": "30-80 ตร.ม.",
        "sale_price": "3,400,000 - 10,000,000",
        "rent_price": "20,000 - 45,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "BTS หมอชิต / MRT จตุจักร",
        "pet_policy": "ไม่เกิน 15 กก. ต่อ 50 ตร.ม."
    },
    {
        "name": "Equinox Phahol-Vipha",
        "name_th": "อิควินอกซ์ พหล-วิภา",
        "district": "เขตจตุจักร",
        "address": "ถนนพหลโยธิน แขวงจตุจักร เขตจตุจักร กรุงเทพฯ",
        "lat": 13.8120, "lon": 100.5580,
        "units": 150,
        "room_sizes": "30-65 ตร.ม.",
        "sale_price": "4,000,000 - 10,000,000",
        "rent_price": "25,000 - 50,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS หมอชิต",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },
    {
        "name": "M Ladprao",
        "name_th": "เอ็ม ลาดพร้าว",
        "district": "เขตจตุจักร",
        "address": "ถนนลาดพร้าว แขวงจอมพล เขตจตุจักร กรุงเทพฯ",
        "lat": 13.8000, "lon": 100.5600,
        "units": 300,
        "room_sizes": "28-60 ตร.ม.",
        "sale_price": "3,500,000 - 8,000,000",
        "rent_price": "20,000 - 40,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT พหลโยธิน",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Metris District Ladprao",
        "name_th": "เมทริส ดิสทริค ลาดพร้าว",
        "district": "เขตจตุจักร",
        "address": "ถนนลาดพร้าว แขวงจอมพล เขตจตุจักร กรุงเทพฯ 10900",
        "lat": 13.8020, "lon": 100.5620,
        "units": 600,
        "room_sizes": "23.7-49.6 ตร.ม.",
        "sale_price": "4,450,000 - 8,000,000",
        "rent_price": "20,000 - 45,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT พหลโยธิน",
        "pet_policy": "อนุญาตสัตว์เลี้ยงทุกชนิด (รวม Exotic Pet)"
    },
    {
        "name": "Modiz Vault Kaset-Sripatum",
        "name_th": "โมดิซ วอลท์ เกษตร-ศรีปทุม",
        "district": "เขตจตุจักร",
        "address": "ซ. พหลโยธิน 49/1 แขวงลาดยาว เขตจตุจักร กรุงเทพฯ 10900",
        "lat": 13.857375, "lon": 100.580815,
        "units": 787,
        "room_sizes": "22-35.5 ตร.ม.",
        "sale_price": "2,290,000 - 4,500,000",
        "rent_price": "8,000 - 18,000",
        "developer": "AssetWise (ASW)",
        "phone": "02-168-0000",
        "email": "info@assetwise.co.th",
        "bts": "BTS บางบัว",
        "pet_policy": "อาคาร B Pet Friendly แยกต่างหาก, มีห้อง Pet Friendly, Pet Play House"
    },

    # ===== เขตลาดพร้าว (Lat Phrao) =====
    {
        "name": "Maru Ladprao 15",
        "name_th": "มารุ ลาดพร้าว 15",
        "district": "เขตลาดพร้าว",
        "address": "ซอยลาดพร้าว 15 แขวงจอมพล เขตจตุจักร กรุงเทพฯ",
        "lat": 13.7980, "lon": 100.5550,
        "units": 300,
        "room_sizes": "28-65 ตร.ม.",
        "sale_price": "5,300,000 - 12,000,000",
        "rent_price": "18,000 - 40,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "MRT ลาดพร้าว",
        "pet_policy": "ไม่เกิน 20 กก./ตัว ห้องละไม่เกิน 2 ตัว"
    },
    {
        "name": "Chewathai Hallmark Ladprao",
        "name_th": "ชีวาทัย ฮอลล์มาร์ค ลาดพร้าว",
        "district": "เขตลาดพร้าว",
        "address": "ซอยโชคชัย 4 แขวงลาดพร้าว เขตลาดพร้าว กรุงเทพฯ",
        "lat": 13.7800, "lon": 100.5700,
        "units": 250,
        "room_sizes": "25-50 ตร.ม.",
        "sale_price": "2,000,000 - 4,000,000",
        "rent_price": "8,000 - 18,000",
        "developer": "Chewathai",
        "phone": "02-123-4567",
        "email": "info@chewathai.com",
        "bts": "MRT โชคชัย 4 (สายเหลือง)",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก (แจ้งล่วงหน้า)"
    },
    {
        "name": "Happy Condo Ladprao 101",
        "name_th": "แฮปปี้ คอนโด ลาดพร้าว 101",
        "district": "เขตลาดพร้าว",
        "address": "ซอยลาดพร้าว 101 แขวงคลองจั่น เขตบางกะปิ กรุงเทพฯ",
        "lat": 13.7700, "lon": 100.6000,
        "units": 200,
        "room_sizes": "25-50 ตร.ม.",
        "sale_price": "2,300,000 - 4,500,000",
        "rent_price": "8,500 - 18,000",
        "developer": "Infinite Real Estate",
        "phone": "02-123-4567",
        "email": "info@infinite.co.th",
        "bts": "MRT ลาดพร้าว 101 (สายเหลือง)",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Freeisland Ladprao 93",
        "name_th": "ฟรีไอส์แลนด์ ลาดพร้าว 93",
        "district": "เขตลาดพร้าว",
        "address": "ซอยลาดพร้าว 93 แขวงคลองเจ้าคุณสิงห์ เขตวังทองหลาง กรุงเทพฯ",
        "lat": 13.7750, "lon": 100.5900,
        "units": 150,
        "room_sizes": "22-45 ตร.ม.",
        "sale_price": "1,800,000 - 3,500,000",
        "rent_price": "9,000 - 18,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "MRT มหาดไทย (สายเหลือง)",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },

    # ===== เขตบางนา (Bang Na) =====
    {
        "name": "Brixton Pet & Play Sukhumvit 107",
        "name_th": "บริกซ์ตัน เพ็ทแอนด์เพลย์ สุขุมวิท 107",
        "district": "เขตบางนา",
        "address": "ถนนสุขุมวิท 107 แขวงบางนา เขตบางนา กรุงเทพฯ",
        "lat": 13.65674, "lon": 100.60217,
        "units": 237,
        "room_sizes": "26.72-30.12 ตร.ม.",
        "sale_price": "2,050,000 - 2,890,000",
        "rent_price": "11,500 - 16,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "BTS แบริ่ง",
        "pet_policy": "ออกแบบเพื่อสัตว์เลี้ยงโดยเฉพาะ"
    },
    {
        "name": "The Muve Sukhumvit 107",
        "name_th": "เดอะ มูฟ สุขุมวิท 107",
        "district": "เขตบางนา",
        "address": "ถนนสุขุมวิท 107 แขวงบางนา เขตบางนา กรุงเทพฯ",
        "lat": 13.6570, "lon": 100.6030,
        "units": 200,
        "room_sizes": "25-35 ตร.ม.",
        "sale_price": "2,000,000 - 3,500,000",
        "rent_price": "12,000 - 18,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS แบริ่ง",
        "pet_policy": "อนุญาตสัตว์เลี้ยง, มีโซนสัตว์เลี้ยง"
    },
    {
        "name": "Arlo Lasalle",
        "name_th": "อาโล ลาซาล",
        "district": "เขตบางนา",
        "address": "ถนนลาซาล แขวงบางนา เขตบางนา กรุงเทพฯ",
        "lat": 13.6650, "lon": 100.6100,
        "units": 180,
        "room_sizes": "22-40 ตร.ม.",
        "sale_price": "1,800,000 - 3,500,000",
        "rent_price": "12,000 - 20,000",
        "developer": "Ananda",
        "phone": "02-123-4567",
        "email": "info@ananda.co.th",
        "bts": "BTS ลาซาล",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก, มีสวนสัตว์เลี้ยง"
    },
    {
        "name": "Swift Condo Abac Bangna",
        "name_th": "สวิฟท์ คอนโด เอแบค บางนา",
        "district": "เขตบางนา",
        "address": "ถนนบางนา-ตราด แขวงบางนา เขตบางนา กรุงเทพฯ",
        "lat": 13.6600, "lon": 100.6300,
        "units": 150,
        "room_sizes": "22-40 ตร.ม.",
        "sale_price": "1,500,000 - 3,000,000",
        "rent_price": "7,500 - 15,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS สถานีบางนา",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },
    {
        "name": "Origin Plug & Play E22 Station",
        "name_th": "ออริจิ้น ปลั๊ก แอนด์ เพลย์ อี22 สเตชั่น",
        "district": "เขตบางนา",
        "address": "สุขุมวิท แขวงบางนา เขตบางนา กรุงเทพฯ 10260",
        "lat": 13.670, "lon": 100.624,
        "units": 300,
        "room_sizes": "22-35 ตร.ม.",
        "sale_price": "1,990,000 - 4,500,000",
        "rent_price": "9,000 - 18,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "BTS บางนา",
        "pet_policy": "Pet Service Solution ครบวงจร"
    },

    # ===== เขตพระโขนง (Phra Khanong) =====
    {
        "name": "Click Condo Sukhumvit 65",
        "name_th": "คลิก คอนโด สุขุมวิท 65",
        "district": "เขตพระโขนง",
        "address": "ซอยสุขุมวิท 65 แขวงพระโขนงเหนือ เขตวัฒนา กรุงเทพฯ",
        "lat": 13.7150, "lon": 100.5950,
        "units": 150,
        "room_sizes": "24-45 ตร.ม.",
        "sale_price": "2,500,000 - 5,000,000",
        "rent_price": "18,000 - 30,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS พระโขนง",
        "pet_policy": "อนุญาตแมวและสุนัขเล็ก"
    },
    {
        "name": "Seven Place Residences",
        "name_th": "เซเว่น เพลส เรสซิเดนซ์",
        "district": "เขตพระโขนง",
        "address": "ซอยสุขุมวิท 44 แขวงพระโขนง เขตคลองเตย กรุงเทพฯ",
        "lat": 13.7140, "lon": 100.5870,
        "units": 80,
        "room_sizes": "40-100 ตร.ม.",
        "sale_price": "6,000,000 - 15,000,000",
        "rent_price": "30,000 - 65,000",
        "developer": "-",
        "phone": "-",
        "email": "-",
        "bts": "BTS พระโขนง",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "The Line Sukhumvit 101",
        "name_th": "เดอะ ไลน์ สุขุมวิท 101",
        "district": "เขตพระโขนง",
        "address": "สุขุมวิท 101 แขวงบางจาก เขตพระโขนง กรุงเทพฯ 10260",
        "lat": 13.692, "lon": 100.606,
        "units": 350,
        "room_sizes": "30-65 ตร.ม.",
        "sale_price": "7,500,000 - 15,000,000",
        "rent_price": "25,000 - 55,000",
        "developer": "Sansiri / BTS Group",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "BTS บางจาก",
        "pet_policy": "อนุญาตแมวและสุนัขขนาดเล็ก"
    },

    # ===== เขตบางแค (Bang Khae) =====
    {
        "name": "The Origin Bangkae",
        "name_th": "ดิ ออริจิ้น บางแค",
        "district": "เขตบางแค",
        "address": "ถนนเพชรเกษม แขวงบางแค เขตบางแค กรุงเทพฯ",
        "lat": 13.6950, "lon": 100.4000,
        "units": 200,
        "room_sizes": "24-50 ตร.ม.",
        "sale_price": "1,500,000 - 3,500,000",
        "rent_price": "8,000 - 18,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "MRT หลักสอง",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "The MUVE Paw Bangkhae",
        "name_th": "เดอะ มูฟ พอว์ บางแค",
        "district": "เขตบางแค",
        "address": "ถนนเพชรเกษม แขวงบางแค เขตบางแค กรุงเทพฯ",
        "lat": 13.6960, "lon": 100.4050,
        "units": 250,
        "room_sizes": "24-35.25 ตร.ม.",
        "sale_price": "1,490,000 - 3,500,000",
        "rent_price": "9,500 - 18,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "MRT หลักสอง",
        "pet_policy": "Pet-friendly มี Pet Park"
    },
    {
        "name": "Origin Place Phetkasem",
        "name_th": "ออริจิ้น เพลส เพชรเกษม",
        "district": "เขตบางแค",
        "address": "ถนนเพชรเกษม แขวงบางแค เขตบางแค กรุงเทพฯ 10160",
        "lat": 13.701, "lon": 100.395,
        "units": 280,
        "room_sizes": "22-45 ตร.ม.",
        "sale_price": "1,790,000 - 4,500,000",
        "rent_price": "8,000 - 18,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "BTS เพชรเกษม 48",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตบางซื่อ (Bang Sue) =====
    {
        "name": "The Base Wongsawang",
        "name_th": "เดอะ เบส วงศ์สว่าง",
        "district": "เขตบางซื่อ",
        "address": "ถนนวงศ์สว่าง แขวงบางซื่อ เขตบางซื่อ กรุงเทพฯ",
        "lat": 13.8150, "lon": 100.5300,
        "units": 300,
        "room_sizes": "25.75-46 ตร.ม.",
        "sale_price": "2,390,000 - 5,000,000",
        "rent_price": "10,000 - 22,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "MRT วงศ์สว่าง",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตราษฎร์บูรณะ (Rat Burana) =====
    {
        "name": "Nue Riverest Ratburana",
        "name_th": "นิว ริเวอร์เรสต์ ราษฎร์บูรณะ",
        "district": "เขตราษฎร์บูรณะ",
        "address": "ถนนสุขสวัสดิ์ แขวงราษฎร์บูรณะ เขตราษฎร์บูรณะ กรุงเทพฯ",
        "lat": 13.6800, "lon": 100.5100,
        "units": 400,
        "room_sizes": "25-55 ตร.ม.",
        "sale_price": "2,000,000 - 5,000,000",
        "rent_price": "9,000 - 20,000",
        "developer": "Noble Development",
        "phone": "02-123-4567",
        "email": "info@noble.com",
        "bts": "-",
        "pet_policy": "ตึกพิเศษ Pet Allowed Tower"
    },
    {
        "name": "The Key Rama 3-Suksawat",
        "name_th": "เดอะ คีย์ พระราม 3-สุขสวัสดิ์",
        "district": "เขตราษฎร์บูรณะ",
        "address": "ถนนสุขสวัสดิ์ แขวงราษฎร์บูรณะ เขตราษฎร์บูรณะ กรุงเทพฯ",
        "lat": 13.6850, "lon": 100.5150,
        "units": 300,
        "room_sizes": "22.85-64.85 ตร.ม.",
        "sale_price": "1,790,000 - 4,500,000",
        "rent_price": "8,000 - 18,000",
        "developer": "Sansiri",
        "phone": "1685",
        "email": "info@sansiri.com",
        "bts": "-",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Nue Riverest Ratburana",
        "name_th": "นิว ริเวอร์เรสต์ ราษฎร์บูรณะ",
        "district": "เขตราษฎร์บูรณะ",
        "address": "ถนนราษฎร์บูรณะ แขวงราษฎร์บูรณะ เขตราษฎร์บูรณะ กรุงเทพฯ 10140",
        "lat": 13.682, "lon": 100.508,
        "units": 350,
        "room_sizes": "22-45 ตร.ม.",
        "sale_price": "1,990,000 - 4,500,000",
        "rent_price": "8,000 - 20,000",
        "developer": "All Inspire",
        "phone": "02-123-4567",
        "email": "info@allinspire.co.th",
        "bts": "BTS ราษฎร์บูรณะ",
        "pet_policy": "เฉพาะอาคาร Pet Allowed Tower"
    },

    # ===== เขตบางพลัด (Bang Phlat) =====
    {
        "name": "Chapter One Spark Charan",
        "name_th": "แชปเตอร์ วัน สปาร์ค จรัญ",
        "district": "เขตบางพลัด",
        "address": "ถนนจรัญสนิทวงศ์ แขวงบางพลัด เขตบางพลัด กรุงเทพฯ",
        "lat": 13.7800, "lon": 100.5000,
        "units": 500,
        "room_sizes": "25-50 ตร.ม.",
        "sale_price": "2,000,000 - 4,500,000",
        "rent_price": "10,000 - 22,000",
        "developer": "Pruksa",
        "phone": "02-123-4567",
        "email": "info@pruksa.com",
        "bts": "MRT บางพลัด",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตหลักสี่ (Lak Si) =====
    {
        "name": "Origin Place Chaengwattana",
        "name_th": "ออริจิ้น เพลส แจ้งวัฒนะ",
        "district": "เขตหลักสี่",
        "address": "ถนนแจ้งวัฒนะ แขวงทุ่งสองห้อง เขตหลักสี่ กรุงเทพฯ",
        "lat": 13.8850, "lon": 100.5650,
        "units": 200,
        "room_sizes": "22-63 ตร.ม.",
        "sale_price": "1,800,000 - 4,000,000",
        "rent_price": "8,500 - 18,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "MRT สายสีชมพู",
        "pet_policy": "Pet-friendly, ห้อง Duo Space"
    },
    {
        "name": "Supalai Sense Chaengwattana-Laksi",
        "name_th": "ศุภาลัย เซนส์ แจ้งวัฒนะ-หลักสี่",
        "district": "เขตหลักสี่",
        "address": "ถนนแจ้งวัฒนะ แขวงทุ่งสองห้อง เขตหลักสี่ กรุงเทพฯ",
        "lat": 13.8830, "lon": 100.5630,
        "units": 300,
        "room_sizes": "25-55 ตร.ม.",
        "sale_price": "1,650,000 - 4,000,000",
        "rent_price": "8,000 - 18,000",
        "developer": "Supalai",
        "phone": "02-123-4567",
        "email": "info@supalai.com",
        "bts": "SRT หลักสี่ / MRT สายสีชมพู",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตสะพานสูง (Saphan Sung) =====
    {
        "name": "COZI Raminthra-Khubon",
        "name_th": "โคซี่ รามอินทรา-คู้บอน",
        "district": "เขตสะพานสูง",
        "address": "ถนนรามอินทรา แขวงสะพานสูง เขตสะพานสูง กรุงเทพฯ",
        "lat": 13.7800, "lon": 100.6800,
        "units": 250,
        "room_sizes": "24-45 ตร.ม.",
        "sale_price": "1,500,000 - 3,500,000",
        "rent_price": "7,000 - 15,000",
        "developer": "SC Asset",
        "phone": "02-123-4567",
        "email": "info@scasset.com",
        "bts": "-",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    {
        "name": "Landmark @Grand Station",
        "name_th": "แลนด์มาร์ค แอท แกรนด์ สเตชั่น",
        "district": "เขตสะพานสูง",
        "address": "ถนนรามอินทรา แขวงสะพานสูง เขตสะพานสูง กรุงเทพฯ",
        "lat": 13.7850, "lon": 100.6850,
        "units": 988,
        "room_sizes": "38.50-50 ตร.ม.",
        "sale_price": "1,980,000 - 4,500,000",
        "rent_price": "30,000 - 50,000",
        "developer": "Siamis Asset",
        "phone": "02-123-4567",
        "email": "info@siamis.com",
        "bts": "MRT สายสีชมพู (วงแหวนรามอินทรา)",
        "pet_policy": "มีห้อง Pet Lover แยกชั้น 4-5"
    },

    # ===== เขตคันนายาว (Khan Na Yao) =====
    {
        "name": "Origin Plug & Play Ramintra",
        "name_th": "ออริจิ้น ปลั๊ก แอนด์ เพลย์ รามอินทรา",
        "district": "เขตคันนายาว",
        "address": "ถนนรามอินทรา แขวงคันนายาว เขตคันนายาว กรุงเทพฯ",
        "lat": 13.8000, "lon": 100.6600,
        "units": 300,
        "room_sizes": "22-45 ตร.ม.",
        "sale_price": "2,790,000 - 5,000,000",
        "rent_price": "9,500 - 18,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "MRT สายสีชมพู (รามอินทรา กม.9)",
        "pet_policy": "อนุญาตสัตว์เลี้ยง, Duo Space"
    },
    {
        "name": "Noble Create",
        "name_th": "โนเบิล ครีเอท",
        "district": "เขตคันนายาว",
        "address": "ถนนรามอินทรา แขวงคันนายาว เขตคันนายาว กรุงเทพฯ",
        "lat": 13.7950, "lon": 100.6700,
        "units": 1250,
        "room_sizes": "25-60 ตร.ม.",
        "sale_price": "2,390,000 - 6,000,000",
        "rent_price": "10,000 - 25,000",
        "developer": "Noble Development",
        "phone": "02-123-4567",
        "email": "info@noble.com",
        "bts": "-",
        "pet_policy": "Tower F Pet Friendly, มี Pet Park"
    },

    # ===== เขตบางกะปิ (Bang Kapi) =====
    {
        "name": "Chateau In Town Kaset Campus",
        "name_th": "ชาโตว์ อินทาวน์ เกษตร แคมปัส",
        "district": "เขตบางกะปิ",
        "address": "ถนนพหลโยธิน แขวงคลองจั่น เขตบางกะปิ กรุงเทพฯ",
        "lat": 13.8080, "lon": 100.5780,
        "units": 180,
        "room_sizes": "22-45 ตร.ม.",
        "sale_price": "1,500,000 - 3,000,000",
        "rent_price": "7,000 - 15,000",
        "developer": "LPN Development",
        "phone": "02-123-4567",
        "email": "info@lpn.co.th",
        "bts": "BTS เกษตร",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },

    # ===== เขตประเวศ (Prawet) =====
    {
        "name": "Metris Pattanakarn",
        "name_th": "เมทริส พัฒนาการ",
        "district": "เขตประเวศ",
        "address": "ซอยพัฒนาการ 12 แขวงสวนหลวง เขตสวนหลวง กรุงเทพฯ 10250",
        "lat": 13.7300, "lon": 100.6200,
        "units": 300,
        "room_sizes": "29.8-61.9 ตร.ม.",
        "sale_price": "3,000,000 - 7,000,000",
        "rent_price": "15,000 - 35,000",
        "developer": "Major Development",
        "phone": "02-116-1111",
        "email": "info@mde.co.th",
        "bts": "ARL รามคำแหง",
        "pet_policy": "อนุญาตแมวและสุนัข"
    },

    # ===== เขตป้อมปราบศัตรูพ่าย (Pom Prap Sattru Phai) =====
    # ===== เขตพระประแดง (Phra Pradaeng) =====
    {
        "name": "Brompton Pet Friendly Samrong Station",
        "name_th": "บรอมป์ตัน เพ็ท เฟรนด์ลี่ สำโรง สเตชั่น",
        "district": "เขตพระประแดง",
        "address": "สุขุมวิท แขวงสำโรงเหนือ เขตพระประแดง สมุทรปราการ",
        "lat": 13.646, "lon": 100.599,
        "units": 250,
        "room_sizes": "22-35 ตร.ม.",
        "sale_price": "1,790,000 - 3,500,000",
        "rent_price": "8,000 - 16,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "BTS สำโรง",
        "pet_policy": "Pet Friendly Project"
    },
    # ===== เขตพระนคร (Phra Nakhon) =====
    # ===== เขตดุสิต (Dusit) =====
    # ===== เขตหนองจอก (Nong Chok) =====
    # ===== เขตบางเขน (Bang Khen) =====
    # ===== เขตมีนบุรี (Min Buri) =====
    # ===== เขตลาดกระบัง (Lat Krabang) =====
    # ===== เขตยานนาวา (Yan Nawa) =====
    # ===== เขตสัมพันธวงศ์ (Samphanthawong) =====
    # ===== เขตธนบุรี (Thon Buri) =====
    # ===== เขตบางกอกใหญ่ (Bangkok Yai) =====
    # ===== เขตคลองสาน (Khlong San) =====
    {
        "name": "Watermark Chaophraya",
        "name_th": "วอร์เตอร์มาร์ค เจ้าพระยา",
        "district": "เขตคลองสาน",
        "address": "346 ถนนเจริญนคร แขวงคลองต้นไทร เขตคลองสาน กรุงเทพฯ 10600",
        "lat": 13.734, "lon": 100.504,
        "units": 400,
        "room_sizes": "45-150 ตร.ม.",
        "sale_price": "8,600,000 - 30,000,000",
        "rent_price": "35,000 - 100,000",
        "developer": "Country Group Development",
        "phone": "02-123-4567",
        "email": "info@countrygroup.co.th",
        "bts": "BTS กรุงธนบุรี",
        "pet_policy": "พื้นที่สำหรับน้องหมาและน้องแมว"
    },
    # ===== เขตตลิ่งชัน (Taling Chan) =====
    # ===== เขตบางกอกน้อย (Bangkok Noi) =====
    # ===== เขตบางขุนเทียน (Bang Khun Thian) =====
    # ===== เขตภาษีเจริญ (Phasi Charoen) =====
    # ===== เขตหนองแขม (Nong Khaem) =====
    # ===== เขตบึงกุ่ม (Bueng Kum) =====
    # ===== เขตบางคอแหลม (Bang Kho Laem) =====
    {
        "name": "Muniq Charoenkrung",
        "name_th": "มิวนีค เจริญกรุง",
        "district": "เขตบางคอแหลม",
        "address": "1957 ถนนเจริญกรุง แขวงวัดพระยาไกร เขตบางคอแหลม กรุงเทพฯ 10120",
        "lat": 13.7000, "lon": 100.5100,
        "units": 168,
        "room_sizes": "92 ตร.ม.",
        "sale_price": "12,000,000 - 30,000,000",
        "rent_price": "45,000 - 100,000",
        "developer": "Major Development",
        "phone": "1266",
        "email": "info@mde.co.th",
        "bts": "BTS เจริญกรุง",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    # ===== เขตสวนหลวง (Suan Luang) =====
    {
        "name": "Park Court Sukhumvit 77",
        "name_th": "พาร์ค คอร์ท สุขุมวิท 77",
        "district": "เขตสวนหลวง",
        "address": "สุขุมวิท 77 แขวงสวนหลวง เขตสวนหลวง กรุงเทพฯ 10250",
        "lat": 13.710, "lon": 100.612,
        "units": 180,
        "room_sizes": "25-50 ตร.ม.",
        "sale_price": "2,500,000 - 5,500,000",
        "rent_price": "12,000 - 25,000",
        "developer": "Land and Houses",
        "phone": "02-123-4567",
        "email": "info@lh.co.th",
        "bts": "BTS อ่อนนุช",
        "pet_policy": "อนุญาตสัตว์เลี้ยง"
    },
    # ===== เขตจอมทอง (Chom Thong) =====
    # ===== เขตดอนเมือง (Don Mueang) =====
    # ===== เขตสายไหม (Sai Mai) =====
    # ===== เขตคลองสามวา (Khlong Sam Wa) =====
    # ===== เขตทวีวัฒนา (Thawi Watthana) =====
    # ===== เขตทุ่งครุ (Thung Khru) =====
    # ===== เขตบางบอน (Bang Bon) =====
    # ===== เขตวังทองหลาง (Wang Thonglang) =====

    # ===== เขตบางเขน (Bang Khen) =====
    {
        "name": "BRIXTON Pet & Play Phaholyothin 50",
        "name_th": "บริกซ์ตัน เพ็ทแอนด์เพลย์ พหลโยธิน 50",
        "district": "เขตสายไหม",
        "address": "ถนนพหลโยธิน 50 แขวงสายไหม เขตสายไหม กรุงเทพฯ",
        "lat": 13.8900, "lon": 100.6100,
        "units": 200,
        "room_sizes": "25-35 ตร.ม.",
        "sale_price": "1,800,000 - 3,000,000",
        "rent_price": "10,000 - 16,000",
        "developer": "Origin Property",
        "phone": "1498",
        "email": "info@origin.co.th",
        "bts": "BTS สายหยุด",
        "pet_policy": "ออกแบบเพื่อสัตว์เลี้ยงโดยเฉพาะ"
    },
]

def get_rent_range(rent_str):
    if not rent_str or rent_str == "-":
        return None
    rent_str = rent_str.replace(",", "").replace("฿", "").replace(" ", "")
    parts = rent_str.split("-")
    nums = []
    for p in parts:
        p = p.strip()
        if p.endswith("+") or p.endswith("+"):
            p = p[:-1]
        try:
            nums.append(int(p))
        except:
            pass
    if nums:
        return nums[0]
    return None

def get_sale_range(sale_str):
    if not sale_str or sale_str == "-":
        return None
    sale_str = sale_str.replace(",", "").replace("฿", "").replace(" ", "")
    parts = sale_str.split("-")
    nums = []
    for p in parts:
        p = p.strip()
        if p.endswith("+") or p.endswith("+"):
            p = p[:-1]
        try:
            nums.append(int(p))
        except:
            pass
    if nums:
        return nums[0]
    return None

def classify_rent_price(rent_str):
    val = get_rent_range(rent_str)
    if val is None:
        return "N/A"
    if val < 10000:
        return "< 10K"
    elif val < 20000:
        return "10K-20K"
    elif val < 50000:
        return "20K-50K"
    elif val < 100000:
        return "50K-100K"
    else:
        return "> 100K"

def classify_sale_price(sale_str):
    val = get_sale_range(sale_str)
    if val is None:
        return "N/A"
    if val < 2000000:
        return "< 2M"
    elif val < 5000000:
        return "2M-5M"
    elif val < 10000000:
        return "5M-10M"
    elif val < 20000000:
        return "10M-20M"
    else:
        return "> 20M"

# Group condos by district
condos_by_district = {}
for c in condos:
    d = c["district"]
    if d not in condos_by_district:
        condos_by_district[d] = []
    condos_by_district[d].append(c)

# ============================================================
# GENERATE EXCEL
# ============================================================
wb = openpyxl.Workbook()

# Styles
header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_fill_rent = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sub_header_font = Font(name="Arial", bold=True, size=11, color="2F5496")
sub_header_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
price_range_fills = {
    "< 10K": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "10K-20K": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "20K-50K": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "50K-100K": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "> 100K": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "< 2M": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "2M-5M": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "5M-10M": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "10M-20M": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "> 20M": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "N/A": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
}
normal_font = Font(name="Arial", size=10)
title_font = Font(name="Arial", bold=True, size=16, color="2F5496")
district_title_font = Font(name="Arial", bold=True, size=14, color="2F5496")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEADERS = ["ลำดับ", "ชื่อโครงการ (EN)", "ชื่อโครงการ (TH)", "ที่อยู่", "เขต", "ละติจูด", "ลองจิจูด",
           "จำนวนยูนิต", "ขนาดห้อง (ตร.ม.)", "ราคาขาย (บาท)", "ช่วงราคาขาย", "ค่าเช่า (บาท/เดือน)", "ช่วงราคาเช่า",
           "Developer", "เบอร์โทร", "อีเมล์", "BTS/MRT ใกล้เคียง", "นโยบายสัตว์เลี้ยง"]

def write_sheet(ws, title, data_list, is_summary=False):
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # Data
    for row_idx, item in enumerate(data_list, 3):
        vals = [
            row_idx - 2,
            item.get("name", ""),
            item.get("name_th", ""),
            item.get("address", ""),
            item.get("district", ""),
            item.get("lat", ""),
            item.get("lon", ""),
            item.get("units", ""),
            item.get("room_sizes", ""),
            item.get("sale_price", ""),
            classify_sale_price(item.get("sale_price", "")),
            item.get("rent_price", ""),
            classify_rent_price(item.get("rent_price", "")),
            item.get("developer", ""),
            item.get("phone", ""),
            item.get("email", ""),
            item.get("bts", ""),
            item.get("pet_policy", ""),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in (1, 6, 7, 8, 11, 13):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # Color price range columns
            if col_idx == 11:  # sale price range
                fill = price_range_fills.get(val, None)
                if fill:
                    cell.fill = fill
            if col_idx == 13:  # rent price range
                fill = price_range_fills.get(val, None)
                if fill:
                    cell.fill = fill

    # Column widths
    col_widths = [6, 25, 25, 45, 12, 10, 10, 10, 16, 20, 14, 20, 14, 18, 15, 25, 22, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = "A3"

# ====== SHEET 1: All Condos ======
ws_all = wb.active
ws_all.title = "รวมทุกโครงการ"
all_condos_sorted = sorted(condos, key=lambda x: (x["district"], x["name"]))
write_sheet(ws_all, "รวมคอนโด Pet Friendly ในกรุงเทพฯ ทั้งหมด", all_condos_sorted)

# ====== SHEETS by District with price range grouping ======
for district in DISTRICTS_TH:
    if district in condos_by_district and len(condos_by_district[district]) > 0:
        d_condos = condos_by_district[district]
        # Sort by rent price
        d_condos_sorted = sorted(d_condos, key=lambda x: get_rent_range(x.get("rent_price", "")) or 999999999)

        ws = wb.create_sheet(title=district.replace("เขต", ""))
        write_sheet(ws, f"คอนโด Pet Friendly - {district}", d_condos_sorted)

# ====== SHEET: Summary ======
ws_summary = wb.create_sheet(title="สรุปตามเขต")
ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
title_cell = ws_summary.cell(row=1, column=1, value="สรุปจำนวนคอนโด Pet Friendly แยกตามเขต กรุงเทพมหานคร")
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_summary.row_dimensions[1].height = 35

summary_headers = ["ลำดับ", "เขต", "จำนวนโครงการ", 
                   "< 10K", "10K-20K", "20K-50K", "50K-100K", "> 100K",
                   "ต่ำสุด (บาท/เดือน)", "สูงสุด (บาท/เดือน)", "ราคาขายเริ่มต้นต่ำสุด"]
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=2, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws_summary.row_dimensions[2].height = 30

row = 3
total_projects = 0
for dist_idx, (dist_th, dist_en) in enumerate(zip(DISTRICTS_TH, DISTRICTS_EN), 1):
    condos_in_dist = condos_by_district.get(dist_th, [])
    count = len(condos_in_dist)
    total_projects += count

    rent_ranges = {"< 10K": 0, "10K-20K": 0, "20K-50K": 0, "50K-100K": 0, "> 100K": 0}
    min_rent = None
    max_rent = None
    min_sale = None
    for c in condos_in_dist:
        r = classify_rent_price(c.get("rent_price", ""))
        if r in rent_ranges:
            rent_ranges[r] += 1
        rv = get_rent_range(c.get("rent_price", ""))
        if rv:
            if min_rent is None or rv < min_rent:
                min_rent = rv
            if max_rent is None or rv > max_rent:
                max_rent = rv
        sv = get_sale_range(c.get("sale_price", ""))
        if sv:
            if min_sale is None or sv < min_sale:
                min_sale = sv

    vals = [dist_idx, f"{dist_th} ({dist_en})", count,
            rent_ranges["< 10K"], rent_ranges["10K-20K"], rent_ranges["20K-50K"],
            rent_ranges["50K-100K"], rent_ranges["> 100K"],
            f"{min_rent:,}" if min_rent else "-",
            f"{max_rent:,}" if max_rent else "-",
            f"{min_sale:,}" if min_sale else "-"]

    for col_idx, val in enumerate(vals, 1):
        cell = ws_summary.cell(row=row, column=col_idx, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align
        if count > 0:
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    row += 1

# Total row
total_rent_dist = {"< 10K": 0, "10K-20K": 0, "20K-50K": 0, "50K-100K": 0, "> 100K": 0}
for c in condos:
    r = classify_rent_price(c.get("rent_price", ""))
    if r in total_rent_dist:
        total_rent_dist[r] += 1

ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
total_cell = ws_summary.cell(row=row, column=1, value=f"รวมทั้งหมด {total_projects} โครงการ")
total_cell.font = Font(name="Arial", bold=True, size=11)
total_cell.alignment = center_align
total_cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
total_cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")

total_vals = [total_projects,
              total_rent_dist["< 10K"], total_rent_dist["10K-20K"], total_rent_dist["20K-50K"],
              total_rent_dist["50K-100K"], total_rent_dist["> 100K"], "", "", ""]
for col_idx, val in enumerate(total_vals, 3):
    cell = ws_summary.cell(row=row, column=col_idx, value=val)
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.alignment = center_align
    cell.border = thin_border

# Column widths for summary
sum_widths = [6, 35, 10, 8, 10, 10, 12, 8, 16, 16, 18]
for i, w in enumerate(sum_widths, 1):
    ws_summary.column_dimensions[get_column_letter(i)].width = w
ws_summary.freeze_panes = "A3"

# ====== SHEET: Price Range Summary ======
ws_price = wb.create_sheet(title="แบ่งตามช่วงราคาเช่า")
ws_price.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
title_cell = ws_price.cell(row=1, column=1, value="คอนโด Pet Friendly แบ่งตามช่วงราคาเช่า")
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_price.row_dimensions[1].height = 35

price_headers = ["ช่วงราคาเช่า", "จำนวนโครงการ", "รายชื่อโครงการ"]
for col_idx, h in enumerate(price_headers, 1):
    cell = ws_price.cell(row=2, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill_rent
    cell.alignment = center_align
    cell.border = thin_border

price_ranges_order = ["< 10K", "10K-20K", "20K-50K", "50K-100K", "> 100K", "N/A"]
row = 3
for pr in price_ranges_order:
    condos_in_pr = [c for c in condos if classify_rent_price(c.get("rent_price", "")) == pr]
    names = ", ".join([c["name"] for c in condos_in_pr])
    cell1 = ws_price.cell(row=row, column=1, value=pr)
    cell1.font = Font(name="Arial", bold=True, size=11)
    cell1.fill = price_range_fills.get(pr, PatternFill())
    cell1.alignment = center_align
    cell1.border = thin_border
    cell2 = ws_price.cell(row=row, column=2, value=len(condos_in_pr))
    cell2.font = normal_font
    cell2.alignment = center_align
    cell2.border = thin_border
    cell3 = ws_price.cell(row=row, column=3, value=names)
    cell3.font = normal_font
    cell3.alignment = left_align
    cell3.border = thin_border
    ws_price.row_dimensions[row].height = max(30, len(condos_in_pr) * 15)
    row += 1

ws_price.column_dimensions["A"].width = 15
ws_price.column_dimensions["B"].width = 12
ws_price.column_dimensions["C"].width = 100

# Save
output_path = "/tmp/opencode/condo_pet/condo_pet_friendly_bangkok.xlsx"
wb.save(output_path)
print(f"✅ Excel saved to: {output_path}")
print(f"   Total projects: {len(condos)}")
print(f"   Districts with data: {len([d for d in DISTRICTS_TH if d in condos_by_district and len(condos_by_district[d]) > 0])}")
print(f"   Total sheets: {len(wb.sheetnames)}")
print(f"   Sheet names: {wb.sheetnames}")
